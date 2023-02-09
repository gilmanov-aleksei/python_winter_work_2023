#! /usr/bin/python

# Импортируем модуль openpyxl
import openpyxl

# Открываем на чтение рабочую книгу
wb = openpyxl.load_workbook("task_10-1.xlsx")
# Записываем в переменную все страницы рабочей книги
sheet = wb.sheetnames
# Записываем в переменную первую страницу
ws = wb[sheet[0]]
# Выводим на экран название активной страницы
# print(ws.title)
# пустой словарь
d = {}
# Цикл по максимальной строке на странице
for i in range(ws.max_row):
    # Пустой список для записи двух ячеек
    z = []
    # Цикл по максимальной колонке на странице
    for j in range(ws.max_column):
        # Записываем две ячейки из строку в список
        z.append(ws.cell(row=i + 1, column=j + 1).value)
    # Проверяем значения списка со словарём
    if d.get(z[0]):
        # если ключ есть, то прибавляем значение
        d[z[0]] += z[1]
    else:
        # если ключа нет, то создаём новый
        # со значением из списка
        d[z[0]] = z[1]

# Счетчик строк
row = 1
# Счетчик сумм
allsum = 0
# Переключаемся на вторую страницу
ws = wb[sheet[1]]
#  Выводим на экарна название второй страницы
# print(ws.title)
# Цикл по словарю
for k, v in d.items():
    # Записваем фамилии в первый столбик
    ws['A' + str(row)] = k
    # Записываем общую суммы
    # по одной фамилии во второй столбик
    ws['B' + str(row)] = v
    # Складываем суммы для ИТОГОВ
    allsum += v
    # Увеличиваем счетчик строк
    row += 1
# Записываем в первой колонке на нижней строке ИТОГО
ws['A' + str(row)] = 'ИТОГО'
# Записываем во второй колонке общую сумму по всем фамилиям
ws['B' + str(row)] = allsum
# СОхраняем все вычисления в файл
wb.save("task_10-1.xlsx")
