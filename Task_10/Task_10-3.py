#! /usr/bin/python


# Импортируем модуль openpyxl
import openpyxl
import statistics

# Открываем на чтение рабочую книгу
wb = openpyxl.load_workbook("task_10-3.xlsx")
# Записываем в переменную все страницы рабочей книги
sheet = wb.sheetnames
# Записываем в переменную первую страницу
ws = wb[sheet[0]]
# пустой словарь
dct = {}
all_stat = []
# Цикл по максимальной строке на странице
for i in range(ws.max_row):
    # Пустой список для записи двух ячеек
    z = []
    # Цикл по максимальной колонке на странице
    for j in range(ws.max_column):
        # Записываем две ячейки из строку в список
        x = ws.cell(row=i + 1, column=j + 1).value
        # Прверяем значение в ячейке, если есть, то
        if x is not None or x != 0:
            # записываем в список, если нет, пропускаем
            z.append(x)
    # Берем из списка имя и удаляем его из списка
    y = z.pop(0)
    # имя это ключ в словаре, список цифр в значение
    dct[y] = z
# print(dct)
    # Вычисляем минимальное, максимальное згачение,
    # Среднеарфметическое и медиану
for k, v in dct.items():
    # Слияние всех списков в один
    all_stat = all_stat + v
    # Расчёт минимального, максимально, среднеарефметического значения и медиан для каждого
    m = [min(v), max(v), round(statistics.mean(v), 1), statistics.median(v)]
    # Перезаписываем значение текущего ключа на произведённое вычисление.
    dct[k] = m
# Расчёт минимального, максимально, среднеарефметического значения и медиан по всем
all_itog = [min(all_stat), max(all_stat), round(statistics.mean(all_stat), 1), statistics.median(all_stat)]
# print(all_stat)
# print(all_itog)
# Производим сортировку словаря по алфавиту
dct = dict(sorted(dct.items()))
# print(dct)
# Открываем на чтение рабочую книгу
wb = openpyxl.load_workbook("task_10-3.xlsx")
# Прверяем в книги, есть ли Лист2,
if 'Лист2' in wb.sheetnames:
    # если есть удаляем его
    wb.remove(wb["Лист2"])
# Создаём второй лист в рабочей книге
wb.create_sheet("Лист2")
# Записываем в переменную все страницы рабочей книги
sheets = wb.sheetnames
# Переключаемся на третию страницу
ws = wb[sheets[1]]
# Счетчик строк
row = 1
for k, v in dct.items():
    # print(str(chr(65)) + str(row), "=", k)
    ws[str(chr(65)) + str(row)] = k
    for i, j in enumerate(v, 66):
        # print(str(chr(i)) + str(row), "=", str(j))
        ws[str(chr(i)) + str(row)] = j
    row += 1
# Записываем в первой колонке на нижней строке ИТОГО
ws['A' + str(row)] = 'ИТОГО'
# Записываем во второй колонке общую сумму по всем фамилиям
# ws['B' + str(row)] = itog_all
for i, j in enumerate(all_itog, 66):
    # print(str(chr(i)) + str(row), "=", str(j))
    ws[str(chr(i)) + str(row)] = j
# Сохраняем все вычисления в файл
wb.save("task_10-3.xlsx")
