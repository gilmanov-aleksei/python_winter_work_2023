#! /usr/bin/python

import openpyxl
from random import randint
from russian_names import RussianNames


def gen_fio_sheet(s, fio, bl, el, name_file):
    # Генератор русских фамилий, имён и отчества русскими буквами
    rn = RussianNames(count=fio, patronymic=False, name=False, surname=True)
    wb = openpyxl.load_workbook(name_file)
    sheets = wb.sheetnames
    ws = wb[sheets[s]]
    fio = rn.get_batch()
    for v, k in enumerate(fio, 1):
        ws['A' + str(v)] = k
        b = ord(bl)
        e = ord(el + 1)
    for i in range(b, e):
        for j in range(1, ws.max_row + 1):
            ws[str(chr(i)) + str(j)] = int(randint(0, 9))
    wb.save(name_file)
    return


# Генератор фамилий из N колонок значений,
# первый аргумент - номер листа куда будут
# записываться фамилиии данные,
# второй аргумент - колличество фамилий в колонке А
# третий аргумент - с какой колонки начать вносить цифры
# четвертый аргумент - какой колонкой закончить внесение цифр
# пятый аргумент - это имя файла, для сохранения результата
gen_fio_sheet(0, 15, 'B', 'K', 'task_10-3.xlsx')
