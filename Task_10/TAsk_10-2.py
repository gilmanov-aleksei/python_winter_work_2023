#! /usr/bin/python

# Импортируем модуль openpyxl
import openpyxl


def data_sheet(s):
    # Открываем на чтение рабочую книгу
    wb = openpyxl.load_workbook("task_10-2.xlsx")
    # Записываем в переменную все страницы рабочей книги
    sheet = wb.sheetnames
    # Записываем в переменную страницу
    ws = wb[sheet[s]]
    # пустой словарь
    d = {}
    # Цикл по максимальной строке на странице
    for i in range(ws.max_row):
        # Пустой список для записи ячеек
        z = []
        # Цикл по максимальной колонке на странице
        for j in range(ws.max_column):
            # Записываем ячейки из строки в список
            z.append(ws.cell(row=i + 1, column=j + 1).value)
        # суммируем срез от второго элемент до конца списка
        # сумму записываем во второй элемент списка
        z[1] = sum(z[1:])
        # Проверяем значения списка со словарём
        if d.get(z[0]):
            # если ключ есть, то прибавляем значение
            d[z[0]] += z[1]
        else:
            # если ключа нет, то создаём новый
            # со значением из списка
            d[z[0]] = z[1]
    return d


print(data_sheet(0))
print(data_sheet(1))
ds0 = data_sheet(0)
ds1 = data_sheet(1)
for k, v in ds1.items():
    if ds0.get(k):
        # если ключ есть, то прибавляем значение
        ds0[k] += v
    else:
        # если ключа нет, то создаём новый
        # со значением из списка
        ds0[k] = v
print(ds0)
