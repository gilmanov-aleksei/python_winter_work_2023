#! /usr/bin/python

# Импортируем модуль openpyxl
import openpyxl


def data_sheet(s):
    # Открываем на чтение рабочую книгу
    wb = openpyxl.load_workbook("task_10-2.xlsx")
    # Записываем в переменную все страницы рабочей книги
    sheet = wb.sheetnames
    # Записываем в переменную страницу
    ws = wb[sheet[s]]
    # пустой словарь
    d = {}
    # Цикл по максимальной строке на странице
    for i in range(ws.max_row):
        # Пустой список для записи ячеек
        z = []
        # Цикл по максимальной колонке на странице
        for j in range(ws.max_column):
            # Записываем ячейки из строки в список
            z.append(ws.cell(row=i + 1, column=j + 1).value)
        # суммируем срез от второго элемент до конца списка
        # сумму записываем во второй элемент списка
        z[1] = sum(z[1:])
        # Проверяем значения списка со словарём
        if d.get(z[0]):
            # если ключ есть, то прибавляем значение
            d[z[0]] += z[1]
        else:
            # если ключа нет, то создаём новый
            # со значением из списка
            d[z[0]] = z[1]
    return d


# С первой страницы вся информация в первый словарь
ds0 = data_sheet(0)
# Со второй страницы вся информация во второй словарь
ds1 = data_sheet(1)
# Цикл сравнения словарей
for k, v in ds1.items():
    # Сравниваем ключ второго словаря с первым
    if ds0.get(k):
        # если ключ есть, то прибавляем значение
        ds0[k] += v
    else:
        # если ключа нет, то создаём новый
        # со значением из второго словаря
        ds0[k] = v

# Сортируем словарь по алфавиту
ds0 = dict(sorted(ds0.items()))
# Открываем на чтение рабочую книгу
wb1 = openpyxl.load_workbook("task_10-2.xlsx")
# Прверяем в книги, есть ли Лист3,
if 'Лист3' in wb1.sheetnames:
    # если есть удаляем его
    wb1.remove(wb1["Лист3"])
# Создаём третию страницу в рабочей книге
wb1.create_sheet("Лист3")
# Записываем в переменную все страницы рабочей книги
sheets = wb1.sheetnames
# Переключаемся на третию страницу
ws1 = wb1[sheets[2]]
# Счетчик строк
row = 1
# Цикл по словарю
for k, v in ds0.items():
    # Записваем фамилии в первый столбик
    ws1['A' + str(row)] = k
    # Записываем значение во второй столбик
    ws1['B' + str(row)] = v
    # Увеличиваем счетчик строк
    row += 1
# Сохраняем все вычисления в файл
wb1.save("task_10-2.xlsx")