#! /usr/bin/python

import openpyxl
# from openpyxl import Workbook
s = []
wb = openpyxl.load_workbook('task_10-0.xlsx')
# print(wb.sheetnames)
for i in wb.sheetnames:
    ws = wb[i]
    s.append((ws.title, ws.max_row * ws.max_column))
print(s)
print(sorted(s, key=lambda x: (x[0], x[1])))
print(sorted(s, key=lambda x: (-x[1], x[0])))


