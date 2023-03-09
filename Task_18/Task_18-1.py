#! /usr/bin/python

# Задача 18 (НЕ ЗАКОНЧЕНА)
# Разработать систему учета решения задач учениками курса «Разработчик на Питоне».
# Проблема. Преподаватель каждый урок задает некоторое количество задач
# в качестве домашнего задания, для упрощения можно считать, что одну.
# Каждый ученик решает каждую задачу. Переводит ее статус в решенную.
# Преподаватель проверяет каждую задачу каждого ученика и либо подтверждает ее
# статус как решенную или меняет ее статус как не решенную.
# Вопрос. Как спроектировать систему классов на Питоне для решения задачи учета.
# Разработайте систему классов (Teacher, Pupil, Lesson, Task. Нужен ли класс Группа)
# Разработайте систему атрибутов для каждого состояния каждого объекта.
# Разработайте систему методов для каждого объекта.
# Отчетность? Запросы? Начните с формулировки решаемой задачи спецификации
# или технического задания. Затем спроектируйте классы, атрибуты, методы.
# Протестируйте систему.


import os.path
import openpyxl


class Teacher:
    def __init__(self, name, subject):
        self.name = name
        self.subject = subject
        self.lessons = []

    def add_lesson(self, lesson):
        self.lessons.append(lesson)

    def delete_lesson(self, lesson):
        self.lessons.remove(lesson)

    def view_lessons(self):
        for lesson in self.lessons:
            print(f"Урок {lesson.lesson_number}")

    def view_lesson(self, lesson_number):
        for lesson in self.lessons:
            if lesson.lesson_number == lesson_number:
                return lesson
        print("Урок не найден")
        return None


class Pupil:
    def __init__(self, name):
        self.name = name
        self.tasks = []

    def add_task(self, task):
        self.tasks.append(task)

    def delete_task(self, task):
        self.tasks.remove(task)

    def view_tasks(self):
        for task in self.tasks:
            print(f"Задание {task.number}: {task.status}")

    def view_task(self, number):
        for task in self.tasks:
            if task.number == number:
                return task
        print("Задание не найдено")
        return None


class Lesson:
    def __init__(self, lesson_number):
        self.lesson_number = lesson_number
        self.tasks = []

    def add_task(self, task):
        self.tasks.append(task)

    def delete_task(self, task):
        self.tasks.remove(task)

    def view_tasks(self):
        for task in self.tasks:
            print(f"Задание {task.number}: {task.status}")

    def view_task(self, number):
        for task in self.tasks:
            if task.number == number:
                return task
        print("Задание не найдено")
        return None


class Task:
    def __init__(self, number, description):
        self.number = number
        self.description = description
        self.status = "Не выполнено"


class ExcelFile:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.sheet_data = {}

    def load_file(self):
        if not os.path.isfile(self.file_path):
            print(f"File {self.file_path} does not exist")
            excel_file.save_file(self.file_path)
            return

        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
        except:
            print(f"File {self.file_path} is already opened in another program")
            return

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]

            row_dict = {}
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                       min_col=1, max_col=sheet.max_column):
                cell_dict = {}
                for cell in row:
                    cell_dict[cell.column_letter] = cell.value
                row_dict[row[0].row] = cell_dict

            self.sheet_data[sheet_name] = row_dict

    def save_file(self, file_path=None):
        if file_path is None:
            file_path = self.file_path

        if self.workbook is None:
            self.workbook = openpyxl.Workbook()

        for sheet_name, row_dict in self.sheet_data.items():
            sheet = self.workbook.create_sheet(sheet_name)

            for row_number, cell_dict in row_dict.items():
                for column_letter, value in cell_dict.items():
                    sheet[f"{column_letter}{row_number}"] = value

        self.workbook.save(file_path)


teacher = Teacher("Михаил", "Разработчик на Питоне")
pupil = Pupil("Алексей")

lesson1 = Lesson(1)
lesson2 = Lesson(2)
task1 = Task(1, "Написать программу для вычисления суммы двух чисел")
task2 = Task(2, "Напишите программу для вычисления произведения двух чисел")

lesson1.add_task(task1)
lesson1.add_task(task2)
teacher.add_lesson(lesson1)
teacher.add_lesson(lesson2)

while True:
    user_input = int(input("Введите число: 1 - Учитель, 2 - Ученик, "
                           "3 - Загрузить из файла, 4 - Сохранить в файл, 0 - Выход: "))
    if user_input == 1:
        teacher_input = int(input("Введите число: 1 - посмотреть уроки, 2 - посмотреть урок по номеру, "
                                  "3 - добавить урок, 4 - удалить урок, 0 - вернуться в меню: "))
        if teacher_input == 1:
            teacher.view_lessons()
        elif teacher_input == 2:
            lesson_number = int(input("Введите номер урока для просмотра: "))
            lesson = Teacher.view_lesson(lesson_number)
            if lesson:
                lesson.view_tasks()
        elif teacher_input == 3:
            lesson_number = int(input("Введите номер урока чтобы добавить: "))
            lesson = Lesson(lesson_number)
            teacher.add_lesson(lesson)
        elif teacher_input == 4:
            lesson_number = int(input("Введите номер урока для удаления: "))
            lesson = Teacher.view_lesson(lesson_number)
            if lesson:
                teacher.delete_lesson(lesson)
    elif user_input == 2:
        print("Ученик")
        # pupil_input = int(input("Введите число: 1 - для просмотра заданий, 2 - для просмотра задания по номеру, "
        #           "3 - добавить задание, 4 - удалить задание, 0 - вернуться в меню: "))
        # if pupil_input == 1:
        #     Pupil.view_tasks()
        # elif pupil_input == 2:
        #     task_number = int(
        #         input("Введите номер задания для просмотра: ")
        #     )
        #     task = Pupil.view_task(task_number)
        #     if task:
        #         print(f"Задание {task.number}: {task.description}, {task.status}")
        # elif pupil_input == 3:
        #     task_number = int(
        #         input("Введит номер для добавления задания: ")
        #     )
        #     task_description = input("Введите номер задания для просмотра: ")
        #     task = Lesson.task(task_number, task_description)
        #     Pupil.add_task(task)
        # elif pupil_input == 4:
        #     task_number = int(
        #         input("Введите номер задания для удаления: ")
        #     )
        #     task = Pupil.view_task(task_number)
        #     if task:
        #         Pupil.delete_task(task)
    elif user_input == 3:
        print("Загрузить из файла")
        excel_file = ExcelFile(f"{input('Name File: ')}.xlsx")
        excel_file.load_file()
    elif user_input == 4:
        print("Сохранить в файл")
        excel_file = ExcelFile(f"{input('Name File: ')}.xlsx")
        excel_file.save_file()
    elif user_input == 0:
        break
