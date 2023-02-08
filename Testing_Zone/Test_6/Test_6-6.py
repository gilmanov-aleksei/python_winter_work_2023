#! /usr/bin/python

import openpyxl
# from openpyxl import Workbook
wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active
# for i in range(1, 4):
#     print(i, ws.cell(row = 1, column = 2).value)
for cellobj in ws['A1':'C3']:
    for cell in cellobj:
        print(cell.coordinate, cell.value)
    print('-------END-------')