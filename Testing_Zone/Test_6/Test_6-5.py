#! /usr/bin/python

import openpyxl
# from openpyxl import Workbook
wb = openpyxl.load_workbook("test.xlsx")
print(wb.sheetnames)
ws = wb.active
print(ws)
print(ws.title)
wb.create_sheet("Sheet1")
sheet = wb['Sheet1']
wb.active = sheet
ws['A4'].value = 1234
c = ws['A4']
print(c.coordinate, c.value)
c.value = c.value * 2
print(ws.max_row)
for i in range(ws.max_row):
    for j in range(ws.max_column):
        print(i + 1, j + 1, ws.cell(row = i + 1, column = j + 1).value)

# wb.save('test.xlsx')
# ws = wb.active
# print(ws)
# wb.active = ws
# print(ws.title)
# print(wb.sheetnames)
# wb.create_sheet("Newssheet")
