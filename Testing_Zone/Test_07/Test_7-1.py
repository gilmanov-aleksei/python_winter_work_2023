#! /usr/bin/python

import openpyxl

wb = openpyxl.load_workbook("test_7.xlsx")
for sh in wb.sheetnames:
    ws = wb[sh]
    print(ws.title, '-------')
    for i in range(ws.max_row):
        for j in range(ws.max_column):
            print(i + 1, j + 1, ws.cell(i + 1, j + 1).value)
    ws = wb[sh]
    ws.append([111, "Текст", 333 ])
    ws.append({1: '123', 3: '345'})
wb.save("test_7.xlsx")
