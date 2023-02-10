#! /usr/bin/python

import openpyxl
from random import randint
from russian_names import RussianNames

# Генератор фамилий и числовых значений в таблицу XLSX

def gen_fio_sheet(s, fio, bl, el, name_file):
    # Генератор русских фамилий, имён и отчества русскими буквами
    rn = RussianNames(count=fio, patronymic=False, name=True, surname=True, transliterate=True)
    wb = openpyxl.load_workbook(name_file)
    sheets = wb.sheetnames
    ws = wb[sheets[s]]
    fio = rn.get_batch()
    for v, k in enumerate(fio, 1):
        ws['A' + str(v)] = k
    for i in range(ord(bl), ord(el) + 1):
        for j in range(1, ws.max_row + 1):
            # Генерация цифр для ячеек
            ws[str(chr(i)) + str(j)] = int(randint(0, 9))
    wb.save(name_file)
    return


# первый аргумент - номер листа куда будут записываться фамилиии,
# второй аргумент - колличество фамилий в колонке А
# третий аргумент - с какой колонки начать вносить цифры
# четвертый аргумент - какой колонкой закончить внесение цифр
# пятый аргумент - это имя файла, для сохранения результата
gen_fio_sheet(0, 10, 'B', 'F', 'test_11-2.xlsx')
