#! /usr/bin/python

# Задача 11-2
#  - Дан файл с расширением .csv ,
#  содержащий в каждой строчке следующую информацию:
#  номер, фамилия, имя, компания, зарплата.
#  - Создайте Эксельный файл, в который перенесите эту информацию,
#  предварительно отсортировав этот список по компании, по фамилии и имени.
#  - В конце списка добавьте строку: ИТОГО и суммарное значение всех зарплат.

# Импортируем модуль openpyxl
import csv
import openpyxl

with open('Task_11-2.csv') as f:
    reader = list(csv.reader(f, delimiter=','))
    begin = reader.pop(0)
    reader.sort(key=lambda x: [x[3], x[1], x[2]])

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Values"
ws = wb['Values']

row = 1
for k, v in enumerate(begin):
    ws[str(chr(65 + k) + str(row))] = v
row += 1
itogo = 0
colum = 0
for m in reader:
    itogo += int(''.join(map(str, m[-1:])))
    for i, j in enumerate(m, 65):
        ws[str(chr(i)) + str(row)] = j
    row += 1
colum = i
ws['A' + str(row)] = 'ИТОГО'
ws[str(chr(colum)) + str(row)] = itogo
wb.save("Task_11-2.xlsx")

